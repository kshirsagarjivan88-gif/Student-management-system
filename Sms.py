import openpyxl
import os

file = "student.xlsx"


# Create Excel file
def create_file():
    if not os.path.exists(file):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name", "Class", "Address", "Contact"])
        wb.save(file)


# Menu
def menu():
    print("\n1. New Student")
    print("2. View Data")
    print("3. Delete Student")
    print("4. Exit")


# New Student
def new_student():
    name = input("Enter Name: ")
    cls = input("Enter Class: ")
    address = input("Enter Address: ")
    contact = input("Enter Contact: ")

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    ws.append([name, cls, address, contact])
    wb.save(file)

    print("Data stored in Excel successfully!")


# View Data
def view_data():
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    print("\n--- Student Data ---")

    for row in ws.iter_rows(values_only=True):
        print(row)


# Delete Student
def delete_student():
    name = input("Enter Name to delete: ")

    wb = openpyxl.load_workbook(file)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == name:
            ws.delete_rows(row)
            wb.save(file)
            print("Student deleted successfully!")
            return

    print("Student not found!")


# Start program
create_file()

while True:
    menu()

    choice = input("Enter choice: ")

    if choice == "1":
        new_student()

    elif choice == "2":
        view_data()

    elif choice == "3":
        delete_student()

    elif choice == "4":
        print("Thank you!")
        break

    else:
        print("Invalid choice!")
